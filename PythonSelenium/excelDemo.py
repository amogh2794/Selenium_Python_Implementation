import openpyxl
file = openpyxl.load_workbook("C:\\PythonUdemyTesting\\PythonSelenium\\pythondemo.xlsx")
#sheet = file.active
#cell = sheet.cell(row=1, column=2)
#data = cell.value
#print(data)
#print(sheet.max_row)
#print(sheet.max_column)

#sheet.cell(row=4, column=2).value = "mambo"

#for i in range(1,sheet.max_row+1):
#    if sheet.cell(row=i, column=1).value == "Testcase2":
#        for j in range(2,sheet.max_column+1):
#            print(sheet.cell(row=i, column=j).value)




#file = openpyxl.load_workbook("C:\\PythonFramework\\excelDemo.py")
sheet = file.active
Dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)